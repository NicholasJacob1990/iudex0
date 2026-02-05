"""
Corpus Service — Camada unificada de gestão da base de conhecimento RAG.

Agrega dados de:
- PostgreSQL (documents table) — metadados e status
- OpenSearch — índices lexicais (BM25)
- Qdrant — coleções vetoriais
- Neo4j — grafo de conhecimento

Inspirado no conceito de "Vault" da Harvey AI.
"""

from __future__ import annotations

import asyncio
import csv
import io
import logging
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import func, case as sql_case, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select

from app.models.document import Document, DocumentCategory, DocumentStatus
from app.models.user import User
from app.schemas.corpus import (
    CorpusAdminActivity,
    CorpusAdminActivityList,
    CorpusAdminOverview,
    CorpusAdminUserList,
    CorpusAdminUserStats,
    CorpusBackfillJurisdictionResponse,
    CorpusBackfillSourceIdResponse,
    CorpusCollectionInfo,
    CorpusDocument,
    CorpusDocumentList,
    CorpusExtendTTLResponse,
    CorpusIngestResponse,
    CorpusPromoteResponse,
    CorpusRetentionPolicy,
    CorpusRetentionPolicyList,
    CorpusSearchResponse,
    CorpusSearchResult,
    CorpusStats,
    CorpusTransferResponse,
)

logger = logging.getLogger(__name__)

# Mapeamento de coleções para display
COLLECTION_DISPLAY = {
    "lei": {"display_name": "Legislação", "description": "Leis, decretos e resoluções"},
    "juris": {"display_name": "Jurisprudência", "description": "Decisões judiciais e precedentes"},
    "pecas_modelo": {"display_name": "Peças Modelo", "description": "Modelos de peças processuais"},
    "doutrina": {"display_name": "Doutrina", "description": "Obras doutrinárias e artigos jurídicos"},
    "sei": {"display_name": "SEI", "description": "Documentos internos (pareceres, notas técnicas)"},
    "local": {"display_name": "Local", "description": "Documentos temporários de casos"},
}

# Mapeamento de collection -> opensearch index name
COLLECTION_TO_OS_INDEX = {
    "lei": "rag-lei",
    "juris": "rag-juris",
    "pecas_modelo": "rag-pecas_modelo",
    "doutrina": "rag-doutrina",
    "sei": "rag-sei",
    "local": "rag-local",
}

# Mapeamento de collection -> qdrant collection name
COLLECTION_TO_QDRANT = {
    "lei": "lei",
    "juris": "juris",
    "pecas_modelo": "pecas_modelo",
    "doutrina": "doutrina",
    "sei": "sei",
    "local": "local_chunks",
}


def _get_opensearch_service():
    """Lazy import do OpenSearch service."""
    from app.services.rag.storage.opensearch_service import get_opensearch_service
    return get_opensearch_service()


def _get_qdrant_service():
    """Lazy import do Qdrant service."""
    from app.services.rag.storage.qdrant_service import get_qdrant_service
    return get_qdrant_service()


def _get_rag_config():
    """Lazy import da configuração RAG."""
    from app.services.rag.config import get_rag_config
    return get_rag_config()


class CorpusService:
    """
    Serviço de gestão unificada do Corpus (base RAG).

    Agrega dados de múltiplos backends para fornecer uma interface
    consistente de consulta e gerenciamento.
    """

    def __init__(self, db: AsyncSession):
        self.db = db

    # =========================================================================
    # Stats
    # =========================================================================

    async def get_stats(
        self,
        user_id: str,
        org_id: Optional[str] = None,
    ) -> CorpusStats:
        """
        Estatísticas agregadas do Corpus.

        Combina contagens do PostgreSQL (documentos) com dados dos backends RAG.
        """
        # Filtro base: documentos do usuário ou da organização
        base_filter = self._build_doc_filter(user_id, org_id)

        # Total de documentos com RAG ingerido
        total_q = select(func.count(Document.id)).where(
            and_(*base_filter, Document.rag_ingested == True)  # noqa: E712
        )
        total_result = await self.db.execute(total_q)
        total_documents = total_result.scalar() or 0

        # Por escopo
        scope_q = (
            select(Document.rag_scope, func.count(Document.id))
            .where(and_(*base_filter, Document.rag_ingested == True))  # noqa: E712
            .group_by(Document.rag_scope)
        )
        scope_result = await self.db.execute(scope_q)
        by_scope = {row[0] or "unknown": row[1] for row in scope_result.all()}

        # Pendentes (status READY mas não ingeridos)
        pending_q = select(func.count(Document.id)).where(
            and_(
                *base_filter,
                Document.rag_ingested == False,  # noqa: E712
                Document.status == DocumentStatus.READY,
            )
        )
        pending_result = await self.db.execute(pending_q)
        pending_ingestion = pending_result.scalar() or 0

        # Falhas (status ERROR)
        failed_q = select(func.count(Document.id)).where(
            and_(*base_filter, Document.status == DocumentStatus.ERROR)
        )
        failed_result = await self.db.execute(failed_q)
        failed_ingestion = failed_result.scalar() or 0

        # Última indexação
        last_q = (
            select(func.max(Document.rag_ingested_at))
            .where(and_(*base_filter, Document.rag_ingested == True))  # noqa: E712
        )
        last_result = await self.db.execute(last_q)
        last_indexed_at = last_result.scalar()

        # Contagens por coleção via backends RAG (best-effort, sem bloquear)
        by_collection = await self._get_collection_counts()

        return CorpusStats(
            total_documents=total_documents,
            by_scope=by_scope,
            by_collection=by_collection,
            pending_ingestion=pending_ingestion,
            failed_ingestion=failed_ingestion,
            last_indexed_at=last_indexed_at,
            storage_size_mb=None,  # Pode ser calculado futuramente
        )

    # =========================================================================
    # List Documents
    # =========================================================================

    async def list_documents(
        self,
        user_id: str,
        org_id: Optional[str] = None,
        scope: Optional[str] = None,
        group_id: Optional[str] = None,
        allowed_group_ids: Optional[List[str]] = None,
        collection: Optional[str] = None,
        search: Optional[str] = None,
        status: Optional[str] = None,
        page: int = 1,
        per_page: int = 20,
    ) -> CorpusDocumentList:
        """
        Lista documentos com informações de ingestão RAG.
        """
        base_filter = self._build_doc_filter(user_id, org_id)
        filters = list(base_filter)

        # Filtro por escopo
        if scope:
            filters.append(Document.rag_scope == scope)

        # Busca por nome
        if search and str(search).strip():
            q = f"%{str(search).strip()}%"
            filters.append(or_(Document.name.ilike(q), Document.original_name.ilike(q)))

        # Filtro por coleção (best-effort: doc_metadata.source_type/dataset ou categoria)
        if collection and str(collection).strip():
            coll = str(collection).strip().lower()
            meta_source_type = Document.doc_metadata["source_type"].as_string()
            meta_dataset = Document.doc_metadata["dataset"].as_string()
            meta_expr = or_(meta_source_type == coll, meta_dataset == coll)

            category_expr = None
            category_map: Dict[str, List[DocumentCategory]] = {
                "lei": [DocumentCategory.LEI],
                "juris": [DocumentCategory.SENTENCA, DocumentCategory.ACORDAO],
                "pecas_modelo": [DocumentCategory.PETICAO],
                "sei": [DocumentCategory.PARECER],
            }
            mapped = category_map.get(coll)
            if mapped:
                category_expr = Document.category.in_(mapped)

            if category_expr is not None:
                filters.append(or_(meta_expr, category_expr))
            else:
                filters.append(meta_expr)

        # Filtro por status
        if status == "ingested":
            filters.append(Document.rag_ingested == True)  # noqa: E712
        elif status == "pending":
            filters.append(Document.rag_ingested == False)  # noqa: E712
            filters.append(Document.status == DocumentStatus.READY)
        elif status == "processing":
            filters.append(Document.status == DocumentStatus.PROCESSING)
        elif status == "failed":
            filters.append(Document.status == DocumentStatus.ERROR)

        normalized_scope = str(scope or "").strip().lower()
        allowed_set = set(str(g).strip() for g in (allowed_group_ids or []) if str(g).strip())
        requested_group_id = str(group_id).strip() if group_id is not None else ""

        offset = (page - 1) * per_page
        documents = []
        total = 0

        # For group scope, apply membership-based filtering using stored doc_metadata.group_ids.
        # This is intentionally defensive (do not leak cross-team) and keeps DB coupling minimal.
        if normalized_scope == "group":
            docs_q = (
                select(Document)
                .where(and_(*filters))
                .order_by(Document.updated_at.desc())
            )
            docs_result = await self.db.execute(docs_q)
            all_docs = docs_result.scalars().all()

            def _get_doc_groups(doc: Document) -> List[str]:
                meta = getattr(doc, "doc_metadata", None) or {}
                raw = meta.get("group_ids") if isinstance(meta, dict) else None
                if not isinstance(raw, list):
                    return []
                return [str(g).strip() for g in raw if str(g).strip()]

            filtered = []
            for doc in all_docs:
                doc_groups = set(_get_doc_groups(doc))
                if allowed_set and doc_groups and not doc_groups.intersection(allowed_set):
                    continue
                if allowed_set and not doc_groups:
                    continue
                if requested_group_id and requested_group_id not in doc_groups:
                    continue
                filtered.append(doc)

            total = len(filtered)
            documents = filtered[offset: offset + per_page]
        else:
            # Contagem total
            count_q = select(func.count(Document.id)).where(and_(*filters))
            count_result = await self.db.execute(count_q)
            total = count_result.scalar() or 0

            # Query paginada
            docs_q = (
                select(Document)
                .where(and_(*filters))
                .order_by(Document.updated_at.desc())
                .offset(offset)
                .limit(per_page)
            )
            docs_result = await self.db.execute(docs_q)
            documents = docs_result.scalars().all()

        items = []
        config = _get_rag_config()
        for doc in documents:
            # Determinar status no Corpus
            if doc.rag_ingested:
                doc_status = "ingested"
            elif doc.status == DocumentStatus.ERROR:
                doc_status = "failed"
            elif doc.status == DocumentStatus.PROCESSING:
                doc_status = "processing"
            else:
                doc_status = "pending"

            # Calcular expiração para documentos locais
            expires_at = None
            if doc.rag_scope == "local" and doc.rag_ingested_at:
                expires_at = doc.rag_ingested_at + timedelta(
                    days=config.local_ttl_days
                )

            items.append(
                CorpusDocument(
                    id=doc.id,
                    name=doc.name,
                    collection=self._infer_collection(doc),
                    scope=doc.rag_scope,
                    status=doc_status,
                    ingested_at=doc.rag_ingested_at,
                    expires_at=expires_at,
                    chunk_count=None,  # Requer consulta ao backend
                    file_type=doc.type.value if doc.type else None,
                    size_bytes=doc.size,
                    jurisdiction=(
                        str((getattr(doc, "doc_metadata", None) or {}).get("jurisdiction") or (getattr(doc, "doc_metadata", None) or {}).get("jurisdição") or "")
                        .strip()
                        .upper()
                        or None
                    ),
                    source_id=(
                        str((getattr(doc, "doc_metadata", None) or {}).get("source_id") or "")
                        .strip()
                        or None
                    ),
                )
            )

        return CorpusDocumentList(
            items=items,
            total=total,
            page=page,
            per_page=per_page,
        )

    # =========================================================================
    # Export Documents
    # =========================================================================

    async def export_documents(
        self,
        *,
        user_id: str,
        org_id: Optional[str] = None,
        scope: Optional[str] = None,
        group_id: Optional[str] = None,
        allowed_group_ids: Optional[List[str]] = None,
        collection: Optional[str] = None,
        search: Optional[str] = None,
        status: Optional[str] = None,
        columns: Optional[List[str]] = None,
        format: str = "csv",
    ) -> Tuple[bytes, str, str]:
        """
        Exporta inventário do Corpus como CSV/XLSX.

        Retorna (content_bytes, filename, content_type).
        """
        fmt = str(format or "csv").strip().lower()
        if fmt not in ("csv", "xlsx"):
            raise ValueError("format inválido. Use 'csv' ou 'xlsx'.")

        allowed_columns = [
            "id",
            "name",
            "original_name",
            "collection",
            "scope",
            "status",
            "ingested_at",
            "expires_at",
            "file_type",
            "size_bytes",
            "case_id",
            "group_ids",
            "jurisdiction",
            "source_id",
        ]

        selected = [c.strip() for c in (columns or []) if str(c).strip()]
        if not selected:
            selected = [
                "name",
                "collection",
                "scope",
                "status",
                "ingested_at",
                "file_type",
                "size_bytes",
            ]

        invalid = [c for c in selected if c not in allowed_columns]
        if invalid:
            raise ValueError(f"Colunas inválidas: {invalid}. Permitidas: {allowed_columns}")

        # Buscar todos os documentos filtrados (sem paginação)
        docs = await self._fetch_documents_for_export(
            user_id=user_id,
            org_id=org_id,
            scope=scope,
            group_id=group_id,
            allowed_group_ids=allowed_group_ids,
            collection=collection,
            search=search,
            status=status,
        )

        config = _get_rag_config()

        def _get_doc_groups(doc: Document) -> List[str]:
            meta = getattr(doc, "doc_metadata", None) or {}
            raw = meta.get("group_ids") if isinstance(meta, dict) else None
            if not isinstance(raw, list):
                return []
            return [str(g).strip() for g in raw if str(g).strip()]

        def _doc_status(doc: Document) -> str:
            if doc.rag_ingested:
                return "ingested"
            if doc.status == DocumentStatus.ERROR:
                return "failed"
            if doc.status == DocumentStatus.PROCESSING:
                return "processing"
            return "pending"

        def _expires_at(doc: Document) -> Optional[datetime]:
            if doc.rag_scope == "local" and doc.rag_ingested_at:
                return doc.rag_ingested_at + timedelta(days=config.local_ttl_days)
            return None

        def _fmt_dt(dt: Optional[datetime]) -> str:
            if not dt:
                return ""
            if isinstance(dt, datetime):
                return dt.astimezone(timezone.utc).isoformat()
            return str(dt)

        rows: List[Dict[str, Any]] = []
        for doc in docs:
            meta = getattr(doc, "doc_metadata", None) or {}
            jurisdiction = None
            source_id = None
            try:
                jurisdiction = str(meta.get("jurisdiction") or meta.get("jurisdição") or "").strip().upper() or None
            except Exception:
                jurisdiction = None
            try:
                source_id = str(meta.get("source_id") or "").strip() or None
            except Exception:
                source_id = None
            row = {
                "id": doc.id,
                "name": doc.name,
                "original_name": doc.original_name,
                "collection": self._infer_collection(doc) or "",
                "scope": doc.rag_scope or "",
                "status": _doc_status(doc),
                "ingested_at": _fmt_dt(doc.rag_ingested_at),
                "expires_at": _fmt_dt(_expires_at(doc)),
                "file_type": doc.type.value if doc.type else "",
                "size_bytes": doc.size,
                "case_id": doc.case_id or "",
                "group_ids": ",".join(_get_doc_groups(doc)),
                "jurisdiction": jurisdiction or "",
                "source_id": source_id or "",
            }
            rows.append({k: row.get(k, "") for k in selected})

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"corpus_documents_{timestamp}.{fmt}"

        if fmt == "csv":
            content = self._render_csv(rows, selected)
            return content, filename, "text/csv; charset=utf-8"

        content = self._render_xlsx(rows, selected)
        return (
            content,
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    async def _fetch_documents_for_export(
        self,
        *,
        user_id: str,
        org_id: Optional[str],
        scope: Optional[str],
        group_id: Optional[str],
        allowed_group_ids: Optional[List[str]],
        collection: Optional[str],
        search: Optional[str],
        status: Optional[str],
    ) -> List[Document]:
        base_filter = self._build_doc_filter(user_id, org_id)
        filters = list(base_filter)

        if scope:
            filters.append(Document.rag_scope == scope)

        if search and str(search).strip():
            q = f"%{str(search).strip()}%"
            filters.append(or_(Document.name.ilike(q), Document.original_name.ilike(q)))

        if collection and str(collection).strip():
            coll = str(collection).strip().lower()
            meta_source_type = Document.doc_metadata["source_type"].as_string()
            meta_dataset = Document.doc_metadata["dataset"].as_string()
            meta_expr = or_(meta_source_type == coll, meta_dataset == coll)

            category_expr = None
            category_map: Dict[str, List[DocumentCategory]] = {
                "lei": [DocumentCategory.LEI],
                "juris": [DocumentCategory.SENTENCA, DocumentCategory.ACORDAO],
                "pecas_modelo": [DocumentCategory.PETICAO],
                "sei": [DocumentCategory.PARECER],
            }
            mapped = category_map.get(coll)
            if mapped:
                category_expr = Document.category.in_(mapped)
            filters.append(or_(meta_expr, category_expr) if category_expr is not None else meta_expr)

        if status == "ingested":
            filters.append(Document.rag_ingested == True)  # noqa: E712
        elif status == "pending":
            filters.append(Document.rag_ingested == False)  # noqa: E712
            filters.append(Document.status == DocumentStatus.READY)
        elif status == "processing":
            filters.append(Document.status == DocumentStatus.PROCESSING)
        elif status == "failed":
            filters.append(Document.status == DocumentStatus.ERROR)

        normalized_scope = str(scope or "").strip().lower()
        allowed_set = set(str(g).strip() for g in (allowed_group_ids or []) if str(g).strip())
        requested_group_id = str(group_id).strip() if group_id is not None else ""

        docs_q = select(Document).where(and_(*filters)).order_by(Document.updated_at.desc())
        docs_result = await self.db.execute(docs_q)
        all_docs = docs_result.scalars().all()

        if normalized_scope != "group":
            return all_docs

        def _get_doc_groups(doc: Document) -> List[str]:
            meta = getattr(doc, "doc_metadata", None) or {}
            raw = meta.get("group_ids") if isinstance(meta, dict) else None
            if not isinstance(raw, list):
                return []
            return [str(g).strip() for g in raw if str(g).strip()]

        filtered: List[Document] = []
        for doc in all_docs:
            doc_groups = set(_get_doc_groups(doc))
            if allowed_set and doc_groups and not doc_groups.intersection(allowed_set):
                continue
            if allowed_set and not doc_groups:
                continue
            if requested_group_id and requested_group_id not in doc_groups:
                continue
            filtered.append(doc)

        return filtered

    def _render_csv(self, rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in columns})
        # UTF-8 BOM para Excel
        return ("\ufeff" + buf.getvalue()).encode("utf-8")

    def _render_xlsx(self, rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Documentos"

        ws.append(columns)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])

        # Ajuste simples de largura
        for idx, col in enumerate(columns, start=1):
            letter = get_column_letter(idx)
            max_len = max(
                [len(str(col or ""))]
                + [len(str(r.get(col, "") or "")) for r in rows[:200]]
            )
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # =========================================================================
    # Ingest
    # =========================================================================

    async def ingest_documents(
        self,
        document_ids: List[str],
        collection: str,
        scope: str,
        user_id: str,
        group_ids: Optional[List[str]] = None,
        jurisdiction: Optional[str] = None,
        source_id: Optional[str] = None,
    ) -> CorpusIngestResponse:
        """
        Dispara ingestão RAG para documentos.

        Utiliza o pipeline de ingestão existente.
        """
        queued = 0
        skipped = 0
        errors: List[Dict[str, str]] = []

        for doc_id in document_ids:
            try:
                # Buscar documento
                result = await self.db.execute(
                    select(Document).where(Document.id == doc_id)
                )
                doc = result.scalar_one_or_none()

                if doc is None:
                    errors.append({
                        "document_id": doc_id,
                        "error": "Documento não encontrado",
                    })
                    continue

                # Verificar se já ingerido
                if doc.rag_ingested and doc.rag_scope == scope:
                    skipped += 1
                    continue

                # Verificar se tem texto extraído
                if not doc.extracted_text and not doc.content:
                    errors.append({
                        "document_id": doc_id,
                        "error": "Documento sem texto extraído. Processe o documento primeiro.",
                    })
                    continue

                text = doc.extracted_text or doc.content or ""

                # Disparar ingestão via pipeline RAG
                try:
                    pipeline = self._get_rag_pipeline()
                    normalized_scope = str(scope or "").strip().lower()
                    normalized_jurisdiction = str(jurisdiction or "").strip().upper() or None
                    if normalized_jurisdiction == "GB":
                        normalized_jurisdiction = "UK"
                    normalized_source_id = str(source_id or "").strip() or None
                    metadata = {
                        "doc_id": doc.id,
                        "tenant_id": doc.organization_id or doc.user_id,
                        "user_id": user_id,
                        "scope": normalized_scope or scope,
                        "source_type": collection,
                        "title": doc.name,
                        "filename": doc.original_name,
                    }
                    if (normalized_scope or scope) == "global":
                        meta_src = doc.doc_metadata if isinstance(doc.doc_metadata, dict) else {}
                        j_from_doc = str(meta_src.get("jurisdiction") or meta_src.get("jurisdição") or "").strip()
                        if j_from_doc:
                            metadata["jurisdiction"] = j_from_doc.strip().upper()
                        elif normalized_jurisdiction:
                            metadata["jurisdiction"] = normalized_jurisdiction
                        else:
                            metadata["jurisdiction"] = "BR"
                        s_from_doc = str(meta_src.get("source_id") or "").strip()
                        if s_from_doc:
                            metadata["source_id"] = s_from_doc
                        elif normalized_source_id:
                            metadata["source_id"] = normalized_source_id
                    if normalized_scope == "group":
                        metadata["group_ids"] = [str(g).strip() for g in (group_ids or []) if str(g).strip()]

                    if scope == "local" and doc.case_id:
                        metadata["case_id"] = doc.case_id

                    if hasattr(pipeline, "ingest_local") and scope == "local":
                        await pipeline.ingest_local(
                            text=text,
                            metadata=metadata,
                            tenant_id=metadata["tenant_id"],
                            case_id=doc.case_id or "default",
                        )
                    elif hasattr(pipeline, "ingest_global") and scope == "global":
                        await pipeline.ingest_global(
                            text=text,
                            metadata=metadata,
                            dataset=collection,
                        )
                    elif hasattr(pipeline, "add_document"):
                        pipeline.add_document(
                            text=text,
                            metadata=metadata,
                            collection=collection,
                        )
                    else:
                        raise RuntimeError(
                            f"Pipeline RAG não suporta ingestão para scope '{scope}'"
                        )

                    # Atualizar status no banco
                    doc.rag_ingested = True
                    doc.rag_ingested_at = datetime.now(timezone.utc)
                    doc.rag_scope = normalized_scope or scope
                    if normalized_scope == "group":
                        meta = doc.doc_metadata or {}
                        meta["group_ids"] = metadata.get("group_ids", [])
                        doc.doc_metadata = meta
                    if metadata.get("jurisdiction"):
                        meta = doc.doc_metadata or {}
                        meta["jurisdiction"] = metadata.get("jurisdiction")
                        doc.doc_metadata = meta
                    if metadata.get("source_id"):
                        meta = doc.doc_metadata or {}
                        meta["source_id"] = metadata.get("source_id")
                        doc.doc_metadata = meta
                    await self.db.commit()
                    queued += 1

                except Exception as e:
                    logger.error(f"Falha na ingestão do documento {doc_id}: {e}")
                    errors.append({
                        "document_id": doc_id,
                        "error": f"Falha na ingestão: {str(e)}",
                    })

            except Exception as e:
                logger.error(f"Erro ao processar documento {doc_id}: {e}")
                errors.append({
                    "document_id": doc_id,
                    "error": str(e),
                })

        return CorpusIngestResponse(
            queued=queued,
            skipped=skipped,
            errors=errors,
        )

    async def backfill_global_jurisdiction(
        self,
        *,
        jurisdiction: str = "BR",
        collections: Optional[List[str]] = None,
        dry_run: bool = True,
        limit: int = 0,
    ) -> CorpusBackfillJurisdictionResponse:
        """
        Backfill do campo de jurisdição nos índices RAG para chunks com `scope=global`.

        - OpenSearch: define `metadata.jurisdiction` quando ausente.
        - Qdrant: define payload `jurisdiction` quando ausente.

        `limit` é um cap de segurança (0 = sem limite).
        """
        normalized = str(jurisdiction or "").strip().upper() or "BR"
        if normalized == "GB":
            normalized = "UK"

        target_collections = [
            str(c).strip()
            for c in (collections or list(COLLECTION_TO_OS_INDEX.keys()))
            if str(c).strip() and str(c).strip() != "local"
        ]
        target_collections = list(dict.fromkeys(target_collections))
        if not target_collections:
            target_collections = [c for c in COLLECTION_TO_OS_INDEX.keys() if c != "local"]

        os_service = None
        qdrant = None
        try:
            os_service = _get_opensearch_service()
        except Exception as exc:
            logger.warning(f"OpenSearch indisponível para backfill: {exc}")
        try:
            qdrant = _get_qdrant_service()
        except Exception as exc:
            logger.warning(f"Qdrant indisponível para backfill: {exc}")

        os_updated_total = 0
        qdrant_updated_total = 0
        os_details: Dict[str, Any] = {}
        qdrant_details: Dict[str, Any] = {}

        # OpenSearch backfill
        if os_service is not None:
            for coll in target_collections:
                index = COLLECTION_TO_OS_INDEX.get(coll)
                if not index:
                    continue

                body: Dict[str, Any] = {
                    "query": {
                        "bool": {
                            "must": [{"term": {"scope": "global"}}],
                            "must_not": [{"exists": {"field": "metadata.jurisdiction"}}],
                        }
                    }
                }
                if limit and limit > 0:
                    body["max_docs"] = int(limit)

                if dry_run:
                    try:
                        count_resp = os_service.client.count(index=index, body=body)
                        os_details[coll] = {"missing": int(count_resp.get("count", 0)), "updated": 0}
                    except Exception as exc:
                        os_details[coll] = {"error": str(exc)}
                    continue

                body["script"] = {
                    "source": (
                        "if (ctx._source.metadata == null) { ctx._source.metadata = [:]; } "
                        "if (ctx._source.metadata.jurisdiction == null) { ctx._source.metadata.jurisdiction = params.j; }"
                    ),
                    "lang": "painless",
                    "params": {"j": normalized},
                }
                try:
                    resp = os_service.client.update_by_query(
                        index=index,
                        body=body,
                        refresh=True,
                        conflicts="proceed",
                        wait_for_completion=True,
                    )
                    updated = int(resp.get("updated", 0) or 0)
                    os_updated_total += updated
                    os_details[coll] = {"updated": updated, "took_ms": resp.get("took")}
                except Exception as exc:
                    os_details[coll] = {"error": str(exc)}

        # Qdrant backfill
        if qdrant is not None:
            try:
                from qdrant_client.http import models as qmodels  # type: ignore
            except Exception:
                qmodels = None

            for coll in target_collections:
                qdrant_collection = COLLECTION_TO_QDRANT.get(coll)
                if not qdrant_collection:
                    continue
                if qmodels is None:
                    qdrant_details[coll] = {"error": "qdrant_client não disponível"}
                    continue

                updated = 0
                scanned = 0
                missing = 0
                next_offset = None
                batch_size = 256

                q_filter = qmodels.Filter(
                    must=[
                        qmodels.FieldCondition(
                            key="scope",
                            match=qmodels.MatchValue(value="global"),
                        )
                    ]
                )

                while True:
                    if limit and scanned >= limit:
                        break
                    page_limit = batch_size if not limit else min(batch_size, max(0, limit - scanned))
                    if page_limit <= 0:
                        break

                    points, next_offset = qdrant.client.scroll(
                        collection_name=qdrant_collection,
                        scroll_filter=q_filter,
                        limit=page_limit,
                        offset=next_offset,
                        with_payload=True,
                        with_vectors=False,
                    )
                    if not points:
                        break
                    scanned += len(points)

                    to_patch: List[Any] = []
                    for p in points:
                        payload = getattr(p, "payload", None) or {}
                        if payload.get("jurisdiction"):
                            continue
                        missing += 1
                        to_patch.append(getattr(p, "id", None))
                    to_patch = [pid for pid in to_patch if pid is not None]

                    if to_patch and not dry_run:
                        qdrant.client.set_payload(
                            collection_name=qdrant_collection,
                            payload={"jurisdiction": normalized},
                            points=to_patch,
                            wait=True,
                        )
                        updated += len(to_patch)

                    if next_offset is None:
                        break

                qdrant_updated_total += updated
                qdrant_details[coll] = {
                    "scanned": scanned,
                    "missing": missing,
                    "updated": 0 if dry_run else updated,
                }

        return CorpusBackfillJurisdictionResponse(
            jurisdiction=normalized,
            dry_run=bool(dry_run),
            collections=target_collections,
            opensearch_updated=os_updated_total,
            qdrant_updated=qdrant_updated_total,
            opensearch_details=os_details,
            qdrant_details=qdrant_details,
        )

    async def backfill_global_source_id(
        self,
        *,
        source_id: str,
        collections: Optional[List[str]] = None,
        dry_run: bool = True,
        limit: int = 0,
    ) -> CorpusBackfillSourceIdResponse:
        """
        Backfill do campo `source_id` nos índices RAG para chunks com `scope=global`.

        - OpenSearch: define `metadata.source_id` quando ausente.
        - Qdrant: define payload `source_id` quando ausente.

        `limit` é um cap de segurança (0 = sem limite).
        """
        normalized = str(source_id or "").strip()
        if not normalized:
            raise ValueError("source_id é obrigatório")

        target_collections = [
            str(c).strip()
            for c in (collections or list(COLLECTION_TO_OS_INDEX.keys()))
            if str(c).strip() and str(c).strip() != "local"
        ]
        target_collections = list(dict.fromkeys(target_collections))
        if not target_collections:
            target_collections = [c for c in COLLECTION_TO_OS_INDEX.keys() if c != "local"]

        os_service = None
        qdrant = None
        try:
            os_service = _get_opensearch_service()
        except Exception as exc:
            logger.warning(f"OpenSearch indisponível para backfill: {exc}")
        try:
            qdrant = _get_qdrant_service()
        except Exception as exc:
            logger.warning(f"Qdrant indisponível para backfill: {exc}")

        os_updated_total = 0
        qdrant_updated_total = 0
        os_details: Dict[str, Any] = {}
        qdrant_details: Dict[str, Any] = {}

        # OpenSearch backfill
        if os_service is not None:
            for coll in target_collections:
                index = COLLECTION_TO_OS_INDEX.get(coll)
                if not index:
                    continue

                body: Dict[str, Any] = {
                    "query": {
                        "bool": {
                            "must": [{"term": {"scope": "global"}}],
                            "must_not": [{"exists": {"field": "metadata.source_id"}}],
                        }
                    }
                }
                if limit and limit > 0:
                    body["max_docs"] = int(limit)

                if dry_run:
                    try:
                        count_resp = os_service.client.count(index=index, body=body)
                        os_details[coll] = {"missing": int(count_resp.get("count", 0)), "updated": 0}
                    except Exception as exc:
                        os_details[coll] = {"error": str(exc)}
                    continue

                body["script"] = {
                    "source": (
                        "if (ctx._source.metadata == null) { ctx._source.metadata = [:]; } "
                        "if (ctx._source.metadata.source_id == null) { ctx._source.metadata.source_id = params.s; }"
                    ),
                    "lang": "painless",
                    "params": {"s": normalized},
                }
                try:
                    resp = os_service.client.update_by_query(
                        index=index,
                        body=body,
                        refresh=True,
                        conflicts="proceed",
                        wait_for_completion=True,
                    )
                    updated = int(resp.get("updated", 0) or 0)
                    os_updated_total += updated
                    os_details[coll] = {"updated": updated, "took_ms": resp.get("took")}
                except Exception as exc:
                    os_details[coll] = {"error": str(exc)}

        # Qdrant backfill
        if qdrant is not None:
            try:
                from qdrant_client.http import models as qmodels  # type: ignore
            except Exception:
                qmodels = None

            for coll in target_collections:
                qdrant_collection = COLLECTION_TO_QDRANT.get(coll)
                if not qdrant_collection:
                    continue
                if qmodels is None:
                    qdrant_details[coll] = {"error": "qdrant_client não disponível"}
                    continue

                updated = 0
                scanned = 0
                missing = 0
                next_offset = None
                batch_size = 256

                q_filter = qmodels.Filter(
                    must=[
                        qmodels.FieldCondition(
                            key="scope",
                            match=qmodels.MatchValue(value="global"),
                        )
                    ]
                )

                while True:
                    if limit and scanned >= limit:
                        break
                    page_limit = batch_size if not limit else min(batch_size, max(0, limit - scanned))
                    if page_limit <= 0:
                        break

                    points, next_offset = qdrant.client.scroll(
                        collection_name=qdrant_collection,
                        scroll_filter=q_filter,
                        limit=page_limit,
                        offset=next_offset,
                        with_payload=True,
                        with_vectors=False,
                    )
                    if not points:
                        break
                    scanned += len(points)

                    to_patch: List[Any] = []
                    for p in points:
                        payload = getattr(p, "payload", None) or {}
                        if payload.get("source_id"):
                            continue
                        missing += 1
                        to_patch.append(getattr(p, "id", None))
                    to_patch = [pid for pid in to_patch if pid is not None]

                    if to_patch and not dry_run:
                        qdrant.client.set_payload(
                            collection_name=qdrant_collection,
                            payload={"source_id": normalized},
                            points=to_patch,
                            wait=True,
                        )
                        updated += len(to_patch)

                    if next_offset is None:
                        break

                qdrant_updated_total += updated
                qdrant_details[coll] = {
                    "scanned": scanned,
                    "missing": missing,
                    "updated": 0 if dry_run else updated,
                }

        return CorpusBackfillSourceIdResponse(
            source_id=normalized,
            dry_run=bool(dry_run),
            collections=target_collections,
            opensearch_updated=os_updated_total,
            qdrant_updated=qdrant_updated_total,
            opensearch_details=os_details,
            qdrant_details=qdrant_details,
        )

    # =========================================================================
    # Remove from Corpus
    # =========================================================================

    async def remove_from_corpus(
        self,
        document_id: str,
        user_id: str,
    ) -> bool:
        """
        Remove documento dos índices RAG (OpenSearch + Qdrant).

        Não exclui o documento do PostgreSQL.
        """
        result = await self.db.execute(
            select(Document).where(Document.id == document_id)
        )
        doc = result.scalar_one_or_none()

        if doc is None:
            return False

        # Verificar permissão
        if doc.user_id != user_id and doc.organization_id is None:
            return False

        deleted_any = False

        # Remover do OpenSearch
        try:
            os_service = _get_opensearch_service()
            count = os_service.delete_by_doc_id(doc_id=document_id)
            if count > 0:
                deleted_any = True
            logger.info(f"OpenSearch: removidos {count} chunks do documento {document_id}")
        except Exception as e:
            logger.error(f"Erro ao remover do OpenSearch: {e}")

        # Remover do Qdrant (tentar todas as coleções)
        try:
            qdrant = _get_qdrant_service()
            for coll_type in COLLECTION_TO_QDRANT.keys():
                try:
                    from app.services.rag.storage.qdrant_service import (
                        Filter,
                        FieldCondition,
                        MatchValue,
                    )
                    delete_filter = Filter(
                        must=[
                            FieldCondition(
                                key="doc_id",
                                match=MatchValue(value=document_id),
                            )
                        ]
                    )
                    qdrant_coll = COLLECTION_TO_QDRANT[coll_type]
                    qdrant.delete_by_filter(
                        collection_type=qdrant_coll,
                        filter_conditions=delete_filter,
                    )
                    deleted_any = True
                except Exception:
                    continue
        except Exception as e:
            logger.error(f"Erro ao remover do Qdrant: {e}")

        # Atualizar status no banco
        doc.rag_ingested = False
        doc.rag_ingested_at = None
        doc.rag_scope = None
        await self.db.commit()

        return True

    # =========================================================================
    # Search
    # =========================================================================

    async def search_corpus(
        self,
        query: str,
        collections: Optional[List[str]] = None,
        scope: Optional[str] = None,
        user_id: str = "",
        org_id: Optional[str] = None,
        limit: int = 10,
    ) -> CorpusSearchResponse:
        """
        Busca unificada no Corpus usando os backends RAG existentes.
        """
        results: List[CorpusSearchResult] = []
        tenant_id = org_id or user_id

        # Determinar índices OpenSearch
        if collections:
            os_indices = [
                COLLECTION_TO_OS_INDEX[c]
                for c in collections
                if c in COLLECTION_TO_OS_INDEX
            ]
        else:
            config = _get_rag_config()
            os_indices = config.get_opensearch_indices()

        # Busca lexical (OpenSearch)
        try:
            os_service = _get_opensearch_service()
            lexical_results = os_service.search_lexical(
                query=query,
                indices=os_indices if os_indices else None,
                top_k=limit,
                scope=scope,
                user_id=user_id,
                include_global=True,
            )
            for hit in lexical_results:
                results.append(
                    CorpusSearchResult(
                        document_id=hit.get("metadata", {}).get("doc_id"),
                        chunk_text=hit.get("text", ""),
                        collection=hit.get("metadata", {}).get("source_type"),
                        score=hit.get("score", 0.0),
                        source="lexical",
                        metadata=hit.get("metadata"),
                    )
                )
        except Exception as e:
            logger.warning(f"Busca lexical falhou: {e}")

        # Busca vetorial (Qdrant) — requer embedding
        try:
            qdrant = _get_qdrant_service()
            embedding = await self._get_embedding(query)

            if embedding:
                qdrant_collections = collections or list(COLLECTION_TO_QDRANT.keys())
                for coll in qdrant_collections:
                    if coll not in COLLECTION_TO_QDRANT:
                        continue
                    try:
                        vector_results = qdrant.search(
                            collection_type=COLLECTION_TO_QDRANT[coll],
                            query_vector=embedding,
                            tenant_id=tenant_id,
                            user_id=user_id,
                            top_k=limit,
                            scopes=[scope] if scope else None,
                        )
                        for hit in vector_results:
                            results.append(
                                CorpusSearchResult(
                                    document_id=hit.metadata.get("doc_id"),
                                    chunk_text=hit.text,
                                    collection=coll,
                                    score=hit.score,
                                    source="vector",
                                    metadata=hit.metadata,
                                )
                            )
                    except Exception as e:
                        logger.warning(f"Busca vetorial em {coll} falhou: {e}")
        except Exception as e:
            logger.warning(f"Busca vetorial falhou: {e}")

        # Ordenar por score (decrescente) e limitar
        results.sort(key=lambda r: r.score, reverse=True)
        results = results[:limit]

        return CorpusSearchResponse(
            results=results,
            total=len(results),
            query=query,
        )

    # =========================================================================
    # Collections
    # =========================================================================

    async def list_collections(self) -> List[CorpusCollectionInfo]:
        """Lista todas as coleções disponíveis no Corpus."""
        collections_info = []

        for name, meta in COLLECTION_DISPLAY.items():
            info = CorpusCollectionInfo(
                name=name,
                display_name=meta["display_name"],
                description=meta["description"],
                document_count=0,
                chunk_count=0,
                scope="global" if name != "local" else "local",
            )

            # Tentar obter contagem do OpenSearch
            try:
                os_service = _get_opensearch_service()
                os_index = COLLECTION_TO_OS_INDEX.get(name)
                if os_index:
                    info.chunk_count = os_service.count_chunks(index=os_index)
            except Exception:
                pass

            # Tentar obter contagem do Qdrant
            try:
                qdrant = _get_qdrant_service()
                qdrant_name = COLLECTION_TO_QDRANT.get(name)
                if qdrant_name:
                    qdrant_info = qdrant.get_collection_info(qdrant_name)
                    if qdrant_info:
                        info.vector_count = qdrant_info.get("points_count")
            except Exception:
                pass

            collections_info.append(info)

        return collections_info

    async def get_collection_info(self, collection: str) -> Optional[CorpusCollectionInfo]:
        """Obtém informações detalhadas de uma coleção."""
        if collection not in COLLECTION_DISPLAY:
            return None

        meta = COLLECTION_DISPLAY[collection]
        info = CorpusCollectionInfo(
            name=collection,
            display_name=meta["display_name"],
            description=meta["description"],
            document_count=0,
            chunk_count=0,
            scope="global" if collection != "local" else "local",
        )

        # OpenSearch
        try:
            os_service = _get_opensearch_service()
            os_index = COLLECTION_TO_OS_INDEX.get(collection)
            if os_index:
                info.chunk_count = os_service.count_chunks(index=os_index)
        except Exception:
            pass

        # Qdrant
        try:
            qdrant = _get_qdrant_service()
            qdrant_name = COLLECTION_TO_QDRANT.get(collection)
            if qdrant_name:
                qdrant_info = qdrant.get_collection_info(qdrant_name)
                if qdrant_info:
                    info.vector_count = qdrant_info.get("points_count")
                    info.status = qdrant_info.get("status", "active")
        except Exception:
            pass

        return info

    # =========================================================================
    # Promote / Extend TTL
    # =========================================================================

    async def promote_local_to_private(
        self,
        document_id: str,
        user_id: str,
    ) -> CorpusPromoteResponse:
        """
        Promove documento de escopo local (temporário) para privado (permanente).
        """
        result = await self.db.execute(
            select(Document).where(Document.id == document_id)
        )
        doc = result.scalar_one_or_none()

        if doc is None:
            return CorpusPromoteResponse(
                document_id=document_id,
                old_scope="unknown",
                new_scope="unknown",
                success=False,
                message="Documento não encontrado.",
            )

        if doc.user_id != user_id:
            return CorpusPromoteResponse(
                document_id=document_id,
                old_scope=doc.rag_scope or "unknown",
                new_scope=doc.rag_scope or "unknown",
                success=False,
                message="Sem permissão para alterar este documento.",
            )

        if doc.rag_scope != "local":
            return CorpusPromoteResponse(
                document_id=document_id,
                old_scope=doc.rag_scope or "unknown",
                new_scope=doc.rag_scope or "unknown",
                success=False,
                message=f"Documento não está no escopo local (escopo atual: {doc.rag_scope}).",
            )

        old_scope = doc.rag_scope
        doc.rag_scope = "private"
        await self.db.commit()

        # TODO: Atualizar metadados nos backends (OpenSearch/Qdrant) para refletir novo escopo

        return CorpusPromoteResponse(
            document_id=document_id,
            old_scope=old_scope or "local",
            new_scope="private",
            success=True,
            message="Documento promovido de local para privado com sucesso.",
        )

    async def extend_local_ttl(
        self,
        document_id: str,
        days: int,
        user_id: str,
    ) -> CorpusExtendTTLResponse:
        """
        Estende o TTL de um documento local.
        """
        result = await self.db.execute(
            select(Document).where(Document.id == document_id)
        )
        doc = result.scalar_one_or_none()

        if doc is None:
            return CorpusExtendTTLResponse(
                document_id=document_id,
                success=False,
                message="Documento não encontrado.",
            )

        if doc.user_id != user_id:
            return CorpusExtendTTLResponse(
                document_id=document_id,
                success=False,
                message="Sem permissão para alterar este documento.",
            )

        if doc.rag_scope != "local":
            return CorpusExtendTTLResponse(
                document_id=document_id,
                success=False,
                message="Apenas documentos locais possuem TTL.",
            )

        # Atualizar rag_ingested_at para simular extensão do TTL
        config = _get_rag_config()
        base_date = doc.rag_ingested_at or datetime.now(timezone.utc)
        new_ingested_at = base_date + timedelta(days=days)

        # Não permitir que ultrapasse o máximo
        max_date = datetime.now(timezone.utc) + timedelta(days=365)
        if new_ingested_at > max_date:
            new_ingested_at = max_date

        doc.rag_ingested_at = new_ingested_at
        await self.db.commit()

        new_expires_at = new_ingested_at + timedelta(days=config.local_ttl_days)

        return CorpusExtendTTLResponse(
            document_id=document_id,
            new_expires_at=new_expires_at,
            success=True,
            message=f"TTL estendido em {days} dias.",
        )

    # =========================================================================
    # Retention Policy
    # =========================================================================

    async def get_retention_policies(
        self,
        org_id: Optional[str] = None,
    ) -> CorpusRetentionPolicyList:
        """
        Retorna as políticas de retenção.

        1. Busca políticas persistidas no banco para a organização.
        2. Preenche escopos faltantes com defaults do RAGConfig.
        """
        from app.models.corpus_retention import CorpusRetentionConfig

        config = _get_rag_config()

        # Defaults estáticos
        defaults = {
            "local": CorpusRetentionPolicy(
                scope="local",
                collection=None,
                retention_days=config.local_ttl_days,
                auto_delete=True,
            ),
            "private": CorpusRetentionPolicy(
                scope="private",
                collection=None,
                retention_days=None,
                auto_delete=False,
            ),
            "global": CorpusRetentionPolicy(
                scope="global",
                collection=None,
                retention_days=None,
                auto_delete=False,
            ),
        }

        # Se temos org_id, buscar políticas persistidas
        if org_id:
            try:
                result = await self.db.execute(
                    select(CorpusRetentionConfig).where(
                        CorpusRetentionConfig.organization_id == org_id
                    )
                )
                db_configs = result.scalars().all()

                for cfg in db_configs:
                    key = cfg.scope
                    defaults[key] = CorpusRetentionPolicy(
                        scope=cfg.scope,
                        collection=cfg.collection,
                        retention_days=cfg.retention_days,
                        auto_delete=cfg.auto_delete,
                    )
            except Exception as e:
                logger.warning(
                    f"Falha ao buscar políticas de retenção do banco para org {org_id}: {e}"
                )
                # Fallback para defaults

        policies = list(defaults.values())
        return CorpusRetentionPolicyList(policies=policies)

    async def update_retention_policy(
        self,
        org_id: str,
        policy: CorpusRetentionPolicy,
    ) -> bool:
        """
        Persiste política de retenção para uma organização.

        Usa upsert: se já existe registro para (org, scope, collection), atualiza;
        caso contrário, cria novo.
        """
        from app.models.corpus_retention import CorpusRetentionConfig

        try:
            # Buscar registro existente
            filters = [
                CorpusRetentionConfig.organization_id == org_id,
                CorpusRetentionConfig.scope == policy.scope,
            ]
            if policy.collection is not None:
                filters.append(CorpusRetentionConfig.collection == policy.collection)
            else:
                filters.append(CorpusRetentionConfig.collection.is_(None))

            result = await self.db.execute(
                select(CorpusRetentionConfig).where(and_(*filters))
            )
            existing = result.scalar_one_or_none()

            if existing:
                existing.retention_days = policy.retention_days
                existing.auto_delete = policy.auto_delete
                logger.info(
                    f"Política de retenção atualizada para org {org_id}: "
                    f"scope={policy.scope}, collection={policy.collection}, "
                    f"retention_days={policy.retention_days}, auto_delete={policy.auto_delete}"
                )
            else:
                import uuid

                new_config = CorpusRetentionConfig(
                    id=str(uuid.uuid4()),
                    organization_id=org_id,
                    scope=policy.scope,
                    collection=policy.collection,
                    retention_days=policy.retention_days,
                    auto_delete=policy.auto_delete,
                )
                self.db.add(new_config)
                logger.info(
                    f"Política de retenção criada para org {org_id}: "
                    f"scope={policy.scope}, collection={policy.collection}, "
                    f"retention_days={policy.retention_days}, auto_delete={policy.auto_delete}"
                )

            await self.db.commit()
            return True

        except Exception as e:
            logger.error(f"Erro ao persistir política de retenção: {e}")
            await self.db.rollback()
            return False

    # =========================================================================
    # Admin — Dashboard administrativo
    # =========================================================================

    async def get_admin_overview(self, org_id: str) -> CorpusAdminOverview:
        """
        Visão geral administrativa: todos os documentos da organização.
        """
        org_filter = [Document.organization_id == org_id]

        # Total de documentos ingeridos
        total_q = select(func.count(Document.id)).where(
            and_(*org_filter, Document.rag_ingested == True)  # noqa: E712
        )
        total_result = await self.db.execute(total_q)
        total_documents = total_result.scalar() or 0

        # Armazenamento total
        storage_q = select(func.coalesce(func.sum(Document.size), 0)).where(
            and_(*org_filter, Document.rag_ingested == True)  # noqa: E712
        )
        storage_result = await self.db.execute(storage_q)
        total_storage_bytes = storage_result.scalar() or 0

        # Usuários ativos (com pelo menos 1 doc ingerido)
        active_q = select(func.count(func.distinct(Document.user_id))).where(
            and_(*org_filter, Document.rag_ingested == True)  # noqa: E712
        )
        active_result = await self.db.execute(active_q)
        active_users = active_result.scalar() or 0

        # Pendentes
        pending_q = select(func.count(Document.id)).where(
            and_(
                *org_filter,
                Document.rag_ingested == False,  # noqa: E712
                Document.status == DocumentStatus.READY,
            )
        )
        pending_result = await self.db.execute(pending_q)
        pending_ingestion = pending_result.scalar() or 0

        # Em processamento
        processing_q = select(func.count(Document.id)).where(
            and_(*org_filter, Document.status == DocumentStatus.PROCESSING)
        )
        processing_result = await self.db.execute(processing_q)
        processing_ingestion = processing_result.scalar() or 0

        # Falhas
        failed_q = select(func.count(Document.id)).where(
            and_(*org_filter, Document.status == DocumentStatus.ERROR)
        )
        failed_result = await self.db.execute(failed_q)
        failed_ingestion = failed_result.scalar() or 0

        # Por escopo
        scope_q = (
            select(Document.rag_scope, func.count(Document.id))
            .where(and_(*org_filter, Document.rag_ingested == True))  # noqa: E712
            .group_by(Document.rag_scope)
        )
        scope_result = await self.db.execute(scope_q)
        by_scope = {row[0] or "unknown": row[1] for row in scope_result.all()}

        # Por coleção
        by_collection = await self._get_collection_counts()

        # Top contribuidores (top 5 por doc_count)
        top_q = (
            select(
                Document.user_id,
                func.count(Document.id).label("doc_count"),
                func.coalesce(func.sum(Document.size), 0).label("storage"),
                func.max(Document.rag_ingested_at).label("last_act"),
            )
            .where(and_(*org_filter, Document.rag_ingested == True))  # noqa: E712
            .group_by(Document.user_id)
            .order_by(func.count(Document.id).desc())
            .limit(5)
        )
        top_result = await self.db.execute(top_q)
        top_rows = top_result.all()

        top_contributors: List[CorpusAdminUserStats] = []
        for row in top_rows:
            user_result = await self.db.execute(
                select(User).where(User.id == row[0])
            )
            user = user_result.scalar_one_or_none()
            top_contributors.append(
                CorpusAdminUserStats(
                    user_id=row[0],
                    user_name=user.name if user else "Desconhecido",
                    user_email=user.email if user else "",
                    doc_count=row[1],
                    storage_bytes=row[2],
                    last_activity=row[3],
                    collections_used=[],
                )
            )

        # Atividade recente (últimos 50 documentos modificados)
        recent_q = (
            select(Document)
            .where(and_(*org_filter))
            .order_by(Document.updated_at.desc())
            .limit(50)
        )
        recent_result = await self.db.execute(recent_q)
        recent_docs = recent_result.scalars().all()

        recent_activity: List[Dict[str, Any]] = []
        for doc in recent_docs:
            action = "ingest"
            if doc.status == DocumentStatus.ERROR:
                action = "failed"
            elif not doc.rag_ingested:
                action = "pending"

            recent_activity.append({
                "document_id": doc.id,
                "document_name": doc.name,
                "user_id": doc.user_id,
                "action": action,
                "timestamp": (doc.rag_ingested_at or doc.updated_at).isoformat()
                if (doc.rag_ingested_at or doc.updated_at)
                else None,
            })

        return CorpusAdminOverview(
            total_documents=total_documents,
            total_storage_bytes=total_storage_bytes,
            active_users=active_users,
            pending_ingestion=pending_ingestion,
            processing_ingestion=processing_ingestion,
            failed_ingestion=failed_ingestion,
            by_collection=by_collection,
            by_scope=by_scope,
            top_contributors=top_contributors,
            recent_activity=recent_activity,
        )

    async def get_corpus_users(
        self,
        org_id: str,
        skip: int = 0,
        limit: int = 20,
    ) -> CorpusAdminUserList:
        """
        Lista usuários da organização com suas estatísticas do Corpus.
        """
        org_filter = [Document.organization_id == org_id]

        # Subquery agrupando por user_id
        user_stats_q = (
            select(
                Document.user_id,
                func.count(Document.id).label("doc_count"),
                func.coalesce(func.sum(Document.size), 0).label("storage"),
                func.max(Document.rag_ingested_at).label("last_act"),
            )
            .where(and_(*org_filter, Document.rag_ingested == True))  # noqa: E712
            .group_by(Document.user_id)
            .order_by(func.count(Document.id).desc())
        )

        # Contagem total de usuários distintos
        count_q = select(func.count(func.distinct(Document.user_id))).where(
            and_(*org_filter, Document.rag_ingested == True)  # noqa: E712
        )
        count_result = await self.db.execute(count_q)
        total = count_result.scalar() or 0

        # Paginar
        paginated_q = user_stats_q.offset(skip).limit(limit)
        result = await self.db.execute(paginated_q)
        rows = result.all()

        items: List[CorpusAdminUserStats] = []
        for row in rows:
            user_result = await self.db.execute(
                select(User).where(User.id == row[0])
            )
            user = user_result.scalar_one_or_none()

            # Coleções usadas pelo usuário
            coll_q = (
                select(func.distinct(Document.rag_scope))
                .where(
                    and_(
                        Document.organization_id == org_id,
                        Document.user_id == row[0],
                        Document.rag_ingested == True,  # noqa: E712
                    )
                )
            )
            coll_result = await self.db.execute(coll_q)
            collections_used = [r[0] for r in coll_result.all() if r[0]]

            items.append(
                CorpusAdminUserStats(
                    user_id=row[0],
                    user_name=user.name if user else "Desconhecido",
                    user_email=user.email if user else "",
                    doc_count=row[1],
                    storage_bytes=row[2],
                    last_activity=row[3],
                    collections_used=collections_used,
                )
            )

        return CorpusAdminUserList(
            items=items,
            total=total,
            skip=skip,
            limit=limit,
        )

    async def get_user_documents(
        self,
        user_id: str,
        org_id: str,
        scope: Optional[str] = None,
        collection: Optional[str] = None,
        status: Optional[str] = None,
        skip: int = 0,
        limit: int = 20,
    ) -> CorpusDocumentList:
        """
        Lista documentos de um usuário específico (visão admin).
        """
        filters = [
            Document.organization_id == org_id,
            Document.user_id == user_id,
        ]

        if scope:
            filters.append(Document.rag_scope == scope)
        if status == "ingested":
            filters.append(Document.rag_ingested == True)  # noqa: E712
        elif status == "pending":
            filters.append(Document.rag_ingested == False)  # noqa: E712
            filters.append(Document.status == DocumentStatus.READY)
        elif status == "processing":
            filters.append(Document.status == DocumentStatus.PROCESSING)
        elif status == "failed":
            filters.append(Document.status == DocumentStatus.ERROR)

        # Contagem
        count_q = select(func.count(Document.id)).where(and_(*filters))
        count_result = await self.db.execute(count_q)
        total = count_result.scalar() or 0

        # Documentos paginados
        docs_q = (
            select(Document)
            .where(and_(*filters))
            .order_by(Document.updated_at.desc())
            .offset(skip)
            .limit(limit)
        )
        docs_result = await self.db.execute(docs_q)
        documents = docs_result.scalars().all()

        config = _get_rag_config()
        items = []
        for doc in documents:
            if doc.rag_ingested:
                doc_status = "ingested"
            elif doc.status == DocumentStatus.ERROR:
                doc_status = "failed"
            elif doc.status == DocumentStatus.PROCESSING:
                doc_status = "processing"
            else:
                doc_status = "pending"

            expires_at = None
            if doc.rag_scope == "local" and doc.rag_ingested_at:
                expires_at = doc.rag_ingested_at + timedelta(
                    days=config.local_ttl_days
                )

            items.append(
                CorpusDocument(
                    id=doc.id,
                    name=doc.name,
                    collection=self._infer_collection(doc),
                    scope=doc.rag_scope,
                    status=doc_status,
                    ingested_at=doc.rag_ingested_at,
                    expires_at=expires_at,
                    chunk_count=None,
                    file_type=doc.type.value if doc.type else None,
                    size_bytes=doc.size,
                )
            )

        page = (skip // limit) + 1 if limit > 0 else 1
        return CorpusDocumentList(
            items=items,
            total=total,
            page=page,
            per_page=limit,
        )

    async def transfer_ownership(
        self,
        document_id: str,
        new_owner_id: str,
        org_id: str,
    ) -> CorpusTransferResponse:
        """
        Transfere a propriedade de um documento para outro usuário da organização.
        """
        # Buscar documento
        result = await self.db.execute(
            select(Document).where(
                Document.id == document_id,
                Document.organization_id == org_id,
            )
        )
        doc = result.scalar_one_or_none()

        if doc is None:
            return CorpusTransferResponse(
                document_id=document_id,
                old_owner_id="",
                new_owner_id=new_owner_id,
                success=False,
                message="Documento não encontrado na organização.",
            )

        # Verificar se novo proprietário existe e pertence à org
        user_result = await self.db.execute(
            select(User).where(
                User.id == new_owner_id,
                User.organization_id == org_id,
            )
        )
        new_owner = user_result.scalar_one_or_none()

        if new_owner is None:
            return CorpusTransferResponse(
                document_id=document_id,
                old_owner_id=doc.user_id,
                new_owner_id=new_owner_id,
                success=False,
                message="Novo proprietário não encontrado na organização.",
            )

        old_owner_id = doc.user_id
        doc.user_id = new_owner_id
        await self.db.commit()

        logger.info(
            f"Documento {document_id} transferido de {old_owner_id} para {new_owner_id} "
            f"(org: {org_id})"
        )

        return CorpusTransferResponse(
            document_id=document_id,
            old_owner_id=old_owner_id,
            new_owner_id=new_owner_id,
            success=True,
            message="Propriedade do documento transferida com sucesso.",
        )

    async def get_corpus_activity(
        self,
        org_id: str,
        skip: int = 0,
        limit: int = 50,
        user_id: Optional[str] = None,
        action: Optional[str] = None,
    ) -> CorpusAdminActivityList:
        """
        Log de atividades do Corpus na organização.
        """
        filters = [Document.organization_id == org_id]

        if user_id:
            filters.append(Document.user_id == user_id)

        if action == "ingested":
            filters.append(Document.rag_ingested == True)  # noqa: E712
        elif action == "failed":
            filters.append(Document.status == DocumentStatus.ERROR)
        elif action == "pending":
            filters.append(Document.rag_ingested == False)  # noqa: E712
            filters.append(Document.status == DocumentStatus.READY)
        elif action == "processing":
            filters.append(Document.status == DocumentStatus.PROCESSING)

        # Contagem total
        count_q = select(func.count(Document.id)).where(and_(*filters))
        count_result = await self.db.execute(count_q)
        total = count_result.scalar() or 0

        # Buscar documentos com dados do usuário
        docs_q = (
            select(Document)
            .where(and_(*filters))
            .order_by(Document.updated_at.desc())
            .offset(skip)
            .limit(limit)
        )
        docs_result = await self.db.execute(docs_q)
        documents = docs_result.scalars().all()

        # Cache de nomes de usuários
        user_cache: Dict[str, str] = {}
        items: List[CorpusAdminActivity] = []

        for doc in documents:
            # Buscar nome do usuário (com cache)
            if doc.user_id not in user_cache:
                u_result = await self.db.execute(
                    select(User.name).where(User.id == doc.user_id)
                )
                u_name = u_result.scalar_one_or_none()
                user_cache[doc.user_id] = u_name or "Desconhecido"

            # Determinar ação
            if doc.rag_ingested:
                doc_action = "ingest"
            elif doc.status == DocumentStatus.ERROR:
                doc_action = "failed"
            elif doc.status == DocumentStatus.PROCESSING:
                doc_action = "processing"
            else:
                doc_action = "pending"

            items.append(
                CorpusAdminActivity(
                    document_id=doc.id,
                    document_name=doc.name,
                    user_id=doc.user_id,
                    user_name=user_cache[doc.user_id],
                    action=doc_action,
                    timestamp=doc.rag_ingested_at or doc.updated_at,
                    details={
                        "scope": doc.rag_scope,
                        "size_bytes": doc.size,
                        "collection": self._infer_collection(doc),
                    },
                )
            )

        return CorpusAdminActivityList(
            items=items,
            total=total,
            skip=skip,
            limit=limit,
        )

    # =========================================================================
    # Helpers privados
    # =========================================================================

    def _build_doc_filter(
        self,
        user_id: str,
        org_id: Optional[str] = None,
    ) -> list:
        """Constrói filtro base para documentos do usuário/organização."""
        if org_id:
            return [Document.organization_id == org_id]
        return [Document.user_id == user_id]

    def _infer_collection(self, doc: Document) -> Optional[str]:
        """Infere a coleção com base nos metadados do documento."""
        meta = doc.doc_metadata or {}
        source_type = meta.get("source_type") or meta.get("dataset")
        if source_type:
            return source_type

        # Inferir pela categoria
        if doc.category:
            category_map = {
                "LEI": "lei",
                "SENTENCA": "juris",
                "ACORDAO": "juris",
                "PETICAO": "pecas_modelo",
                "PARECER": "sei",
            }
            return category_map.get(doc.category.value if hasattr(doc.category, "value") else str(doc.category))

        return None

    async def _get_collection_counts(self) -> Dict[str, int]:
        """Obtém contagens de chunks por coleção dos backends."""
        counts: Dict[str, int] = {}

        try:
            os_service = _get_opensearch_service()
            for name, index in COLLECTION_TO_OS_INDEX.items():
                try:
                    counts[name] = os_service.count_chunks(index=index)
                except Exception:
                    counts[name] = 0
        except Exception:
            pass

        return counts

    async def _get_embedding(self, text: str) -> Optional[List[float]]:
        """Obtém embedding para uma consulta."""
        try:
            from app.services.rag.core.embeddings import get_embeddings_service
            service = get_embeddings_service()
            return service.embed_query(text)
        except ImportError:
            pass
        except Exception as e:
            logger.warning(f"Falha ao obter embedding: {e}")

        # Fallback: tentar via OpenAI diretamente
        try:
            import openai
            config = _get_rag_config()
            client = openai.AsyncOpenAI()
            response = await client.embeddings.create(
                model=config.embedding_model,
                input=text,
            )
            return response.data[0].embedding
        except Exception as e:
            logger.warning(f"Fallback de embedding falhou: {e}")
            return None

    @staticmethod
    def _get_rag_pipeline():
        """Obtém instância do pipeline RAG."""
        try:
            from app.services.rag.pipeline.rag_pipeline import get_rag_pipeline
            return get_rag_pipeline()
        except ImportError:
            try:
                from app.services.rag_module import create_rag_manager
                return create_rag_manager()
            except Exception:
                raise RuntimeError("Pipeline RAG não disponível")
