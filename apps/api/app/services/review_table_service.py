"""
ReviewTableService — Extracao estruturada de dados de documentos em tabelas.

Inspirado no Harvey AI Vault Review Tables:
- Aplica um template de colunas a N documentos
- Para cada documento, extrai cada coluna via chamada de IA
- Armazena resultados em formato tabular (JSON)
- Suporta exportacao como CSV e XLSX
- Column Builder: gera colunas a partir de descricao em linguagem natural
- Fill Table: preenche automaticamente todas as colunas de uma review table
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.time_utils import utcnow
from app.models.document import Document
from app.models.review_table import (
    ColumnType,
    ReviewTable,
    ReviewTableStatus,
    ReviewTableTemplate,
)
from app.services.ai.agent_clients import (
    call_anthropic_async,
    call_vertex_gemini_async,
    get_claude_client,
    get_gemini_client,
)

logger = logging.getLogger("ReviewTableService")

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------

DEFAULT_MODEL = "gemini-3-flash"
MAX_CONCURRENT_EXTRACTIONS = 5
AI_TIMEOUT = 90
MAX_DOC_TEXT_LENGTH = 30000

# Mapeamento completo de tipos de coluna para descricoes de extracao
COLUMN_TYPE_DESCRIPTIONS: Dict[str, str] = {
    "text": "Texto livre (resposta concisa)",
    "date": "Data no formato DD/MM/AAAA",
    "currency": "Valor monetario em reais (formato numerico: 5000.00)",
    "number": "Valor numerico",
    "verbatim": "Transcricao literal do texto do documento",
    "boolean": "Responda apenas 'Sim' ou 'Nao'",
    # Novos tipos — Column Builder
    "summary": "Resumo conciso (2-3 frases) da informacao solicitada",
    "date_extraction": "Data extraida do documento no formato DD/MM/AAAA. Se houver multiplas datas, extraia a mais relevante ao contexto",
    "yes_no_classification": "Classificacao binaria: responda APENAS 'Sim' ou 'Nao' com base na analise do documento",
    "verbatim_extraction": "Transcricao literal e exata do trecho relevante do documento, sem resumo ou interpretacao",
    "risk_rating": "Classificacao de risco: 'Baixo', 'Medio', 'Alto' ou 'Critico'. Justifique brevemente",
    "compliance_check": "Verificacao de conformidade: 'Conforme', 'Nao Conforme' ou 'Parcialmente Conforme'. Indique o fundamento",
    "custom": "Responda conforme a instrucao especifica da coluna",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _safe_json_parse(text: str) -> Any:
    """Tenta parsear JSON mesmo com markdown code fences."""
    if not text:
        return None
    cleaned = text.strip()
    if cleaned.startswith("```"):
        lines = cleaned.split("\n")
        lines = [l for l in lines if not l.strip().startswith("```")]
        cleaned = "\n".join(lines)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        return None


async def _call_ai(
    prompt: str,
    system_instruction: Optional[str] = None,
    model: str = DEFAULT_MODEL,
    max_tokens: int = 2000,
    temperature: float = 0.1,
) -> Optional[str]:
    """Chama modelo de IA com fallback Gemini -> Claude."""
    gemini_client = get_gemini_client()
    if gemini_client:
        try:
            result = await call_vertex_gemini_async(
                client=gemini_client,
                prompt=prompt,
                model=model,
                max_tokens=max_tokens,
                temperature=temperature,
                timeout=AI_TIMEOUT,
                system_instruction=system_instruction,
            )
            if result:
                return result
        except Exception as e:
            logger.warning("Gemini falhou para review table, tentando Claude: %s", e)

    claude_client = get_claude_client()
    if claude_client:
        try:
            result = await call_anthropic_async(
                client=claude_client,
                prompt=prompt,
                model="claude-4.5-sonnet",
                max_tokens=max_tokens,
                temperature=temperature,
                timeout=AI_TIMEOUT,
                system_instruction=system_instruction,
            )
            if result:
                return result
        except Exception as e:
            logger.error("Claude tambem falhou para review table: %s", e)

    return None


EXTRACTION_PROMPT = """Voce e um assistente juridico especializado em extracao de dados de documentos.

Analise o documento abaixo e extraia a informacao solicitada.

## DOCUMENTO
{document_text}

## INFORMACAO A EXTRAIR
{extraction_prompt}

## TIPO DE DADO ESPERADO
{data_type}

## INSTRUCOES
1. Extraia EXATAMENTE a informacao solicitada do documento
2. Se a informacao nao estiver presente no documento, responda: "Nao encontrado"
3. Para tipo "date" ou "date_extraction", use formato DD/MM/AAAA
4. Para tipo "currency", use formato numerico com ponto decimal (ex: 5000.00)
5. Para tipo "boolean" ou "yes_no_classification", responda apenas "Sim" ou "Nao"
6. Para tipo "number", responda apenas com o numero
7. Para tipo "verbatim" ou "verbatim_extraction", transcreva literalmente do documento
8. Para tipo "text" ou "summary", resuma de forma concisa
9. Para tipo "risk_rating", classifique como "Baixo", "Medio", "Alto" ou "Critico"
10. Para tipo "compliance_check", classifique como "Conforme", "Nao Conforme" ou "Parcialmente Conforme"
11. Para tipo "custom", siga a instrucao especifica fornecida

Responda em JSON:
```json
{{
  "value": "<valor extraido>",
  "confidence": <float 0.0 a 1.0>,
  "source_excerpt": "<trecho do documento de onde a informacao foi extraida (max 200 chars)>"
}}
```

Responda APENAS com o JSON, sem texto adicional."""


COLUMN_BUILDER_PROMPT = """Voce e um especialista em due diligence juridica e analise de documentos.

O usuario quer criar uma review table para analisar documentos. Com base na descricao abaixo, gere definicoes de colunas otimizadas para extracao automatica via IA.

## DESCRICAO DO USUARIO
{description}

## TIPOS DE COLUNA DISPONIVEIS
- "summary": Resumo conciso (2-3 frases)
- "date_extraction": Extracao de data (DD/MM/AAAA)
- "yes_no_classification": Classificacao Sim/Nao
- "verbatim_extraction": Transcricao literal do documento
- "risk_rating": Classificacao de risco (Baixo/Medio/Alto/Critico)
- "compliance_check": Verificacao de conformidade (Conforme/Nao Conforme/Parcialmente Conforme)
- "text": Texto livre conciso
- "currency": Valor monetario
- "number": Valor numerico
- "date": Data
- "boolean": Sim/Nao
- "verbatim": Transcricao literal
- "custom": Tipo customizado

## INSTRUCOES
1. Gere entre 3 e 15 colunas relevantes para a descricao
2. Cada coluna DEVE ter um prompt de extracao detalhado e especifico
3. Escolha o tipo mais adequado para cada coluna
4. Os prompts de extracao devem ser claros e direcionados para IA
5. Priorize colunas que agreguem valor na analise juridica
6. Use nomes de coluna concisos em portugues

Responda APENAS em JSON valido:
```json
{{
  "columns": [
    {{
      "name": "<nome da coluna>",
      "type": "<tipo da coluna>",
      "extraction_prompt": "<prompt detalhado para extracao>"
    }}
  ],
  "suggested_name": "<nome sugerido para o template>",
  "suggested_area": "<area juridica: trabalhista, civil, societario, imobiliario, ti, empresarial, tributario, ambiental, regulatorio>"
}}
```

Responda APENAS com o JSON, sem texto adicional."""


# ---------------------------------------------------------------------------
# ReviewTableService
# ---------------------------------------------------------------------------


class ReviewTableService:
    """Servico para extracao estruturada de dados de documentos."""

    # -----------------------------------------------------------------------
    # seed_system_templates
    # -----------------------------------------------------------------------

    async def seed_system_templates(self, db: AsyncSession) -> int:
        """Carrega templates pre-construidos no banco (idempotente)."""
        from app.services.review_table_templates import TEMPLATES

        created = 0
        for tmpl_data in TEMPLATES:
            # Verificar se ja existe template com mesmo nome e is_system=True
            result = await db.execute(
                select(ReviewTableTemplate).where(
                    ReviewTableTemplate.name == tmpl_data["name"],
                    ReviewTableTemplate.is_system == True,  # noqa: E712
                )
            )
            existing = result.scalar_one_or_none()
            if existing:
                # Atualizar colunas se template ja existe
                existing.columns = tmpl_data["columns"]
                existing.description = tmpl_data.get("description")
                existing.area = tmpl_data.get("area")
                existing.updated_at = utcnow()
                continue

            template = ReviewTableTemplate(
                id=str(uuid.uuid4()),
                name=tmpl_data["name"],
                description=tmpl_data.get("description"),
                area=tmpl_data.get("area"),
                columns=tmpl_data["columns"],
                is_system=True,
                created_by=None,
                organization_id=None,
                is_active=True,
            )
            db.add(template)
            created += 1

        await db.commit()
        logger.info("Seed de templates: %d criados", created)
        return created

    # -----------------------------------------------------------------------
    # generate_columns — Column Builder via IA
    # -----------------------------------------------------------------------

    async def generate_columns(
        self,
        description: str,
        area: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Gera definicoes de colunas a partir de uma descricao em linguagem natural.

        Usa IA para interpretar a descricao do usuario e gerar colunas otimizadas
        para extracao automatica, no estilo Harvey AI Column Builder.

        Args:
            description: Descricao em linguagem natural do que o usuario quer analisar.
            area: Area juridica opcional para direcionar a geracao.

        Returns:
            Dict com 'columns' (lista de definicoes), 'suggested_name' e 'suggested_area'.
        """
        if not description or len(description.strip()) < 10:
            raise ValueError(
                "A descricao deve ter no minimo 10 caracteres para gerar colunas."
            )

        enriched_description = description
        if area:
            enriched_description = (
                f"[Area juridica: {area}] {description}"
            )

        prompt = COLUMN_BUILDER_PROMPT.format(description=enriched_description)

        response = await _call_ai(
            prompt=prompt,
            system_instruction=(
                "Voce e um especialista em due diligence e analise juridica. "
                "Gere definicoes de colunas precisas e praticas para review tables. "
                "Responda APENAS em JSON valido."
            ),
            max_tokens=4000,
            temperature=0.3,
        )

        if not response:
            raise RuntimeError(
                "Nao foi possivel gerar colunas. O servico de IA esta indisponivel."
            )

        parsed = _safe_json_parse(response)
        if not parsed or not isinstance(parsed, dict) or "columns" not in parsed:
            raise RuntimeError(
                "Resposta da IA nao pode ser interpretada. Tente reformular a descricao."
            )

        columns = parsed.get("columns", [])

        # Validar e normalizar colunas geradas
        valid_types = {t.value for t in ColumnType}
        validated_columns: List[Dict[str, str]] = []

        for col in columns:
            if not isinstance(col, dict):
                continue
            col_name = col.get("name", "").strip()
            col_type = col.get("type", "text").strip()
            col_prompt = col.get("extraction_prompt", "").strip()

            if not col_name or not col_prompt:
                continue

            # Normalizar tipo
            if col_type not in valid_types:
                col_type = "text"

            validated_columns.append({
                "name": col_name,
                "type": col_type,
                "extraction_prompt": col_prompt,
            })

        if not validated_columns:
            raise RuntimeError(
                "A IA nao gerou colunas validas. Tente reformular a descricao."
            )

        # Limitar a 15 colunas maximo
        validated_columns = validated_columns[:15]

        result = {
            "columns": validated_columns,
            "suggested_name": parsed.get("suggested_name", "Review Table"),
            "suggested_area": parsed.get("suggested_area", area or "geral"),
        }

        logger.info(
            "Column Builder: gerou %d colunas a partir de descricao (%d chars)",
            len(validated_columns),
            len(description),
        )

        return result

    # -----------------------------------------------------------------------
    # create_review
    # -----------------------------------------------------------------------

    async def create_review(
        self,
        template_id: str,
        document_ids: List[str],
        user_id: str,
        org_id: Optional[str],
        name: str,
        db: AsyncSession,
    ) -> ReviewTable:
        """Cria uma nova Review Table e inicia o processamento."""
        # Validar template
        template = await db.get(ReviewTableTemplate, template_id)
        if not template or not template.is_active:
            raise ValueError(f"Template {template_id} nao encontrado ou inativo")

        if not document_ids:
            raise ValueError("E necessario selecionar ao menos um documento")

        if len(document_ids) > 200:
            raise ValueError("Maximo de 200 documentos por review table")

        review = ReviewTable(
            id=str(uuid.uuid4()),
            template_id=template_id,
            name=name,
            user_id=user_id,
            organization_id=org_id,
            status=ReviewTableStatus.CREATED.value,
            document_ids=document_ids,
            results=[],
            total_documents=len(document_ids),
            processed_documents=0,
        )
        db.add(review)
        await db.commit()
        await db.refresh(review)

        logger.info(
            "Review table criada: id=%s, template=%s, docs=%d",
            review.id, template_id, len(document_ids),
        )

        return review

    # -----------------------------------------------------------------------
    # process_review
    # -----------------------------------------------------------------------

    async def process_review(
        self,
        review_id: str,
        db: AsyncSession,
    ) -> ReviewTable:
        """Processa todos os documentos contra as colunas do template."""
        review = await db.get(ReviewTable, review_id)
        if not review:
            raise ValueError(f"Review table {review_id} nao encontrada")

        template = await db.get(ReviewTableTemplate, review.template_id)
        if not template:
            raise ValueError(f"Template {review.template_id} nao encontrado")

        # Marcar como processando
        review.status = ReviewTableStatus.PROCESSING.value
        review.updated_at = utcnow()
        await db.commit()

        columns = template.columns or []
        results: List[Dict[str, Any]] = []
        total_confidence = 0.0
        total_extractions = 0

        semaphore = asyncio.Semaphore(MAX_CONCURRENT_EXTRACTIONS)

        try:
            for doc_id in review.document_ids:
                # Carregar documento
                doc_result = await db.execute(
                    select(Document).where(Document.id == doc_id)
                )
                doc = doc_result.scalar_one_or_none()

                if not doc:
                    results.append({
                        "document_id": doc_id,
                        "document_name": "Documento nao encontrado",
                        "columns": {col["name"]: "Erro: documento nao encontrado" for col in columns},
                    })
                    review.processed_documents += 1
                    continue

                doc_text = doc.extracted_text or doc.content or ""
                if not doc_text.strip():
                    results.append({
                        "document_id": doc_id,
                        "document_name": getattr(doc, "name", doc_id),
                        "columns": {col["name"]: "Erro: sem texto extraido" for col in columns},
                    })
                    review.processed_documents += 1
                    continue

                # Truncar texto
                doc_text_truncated = doc_text[:MAX_DOC_TEXT_LENGTH]
                doc_name = getattr(doc, "name", None) or getattr(doc, "title", None) or doc_id

                # Extrair todas as colunas para este documento
                row_data = await self._extract_row(
                    doc_text=doc_text_truncated,
                    columns=columns,
                    semaphore=semaphore,
                )

                # Calcular confianca
                for col_result in row_data.values():
                    if isinstance(col_result, dict) and "confidence" in col_result:
                        total_confidence += col_result["confidence"]
                        total_extractions += 1

                # Simplificar resultado: apenas valores
                simplified_columns = {}
                column_meta: Dict[str, Any] = {"confidence": {}, "source_excerpt": {}}
                for col_name, col_result in row_data.items():
                    if isinstance(col_result, dict):
                        simplified_columns[col_name] = col_result.get("value", "Nao encontrado")
                        conf = col_result.get("confidence")
                        excerpt = col_result.get("source_excerpt")
                        if conf is not None:
                            try:
                                column_meta["confidence"][col_name] = float(conf)
                            except (TypeError, ValueError):
                                pass
                        if isinstance(excerpt, str) and excerpt.strip():
                            column_meta["source_excerpt"][col_name] = excerpt.strip()
                    else:
                        simplified_columns[col_name] = str(col_result)

                results.append({
                    "document_id": doc_id,
                    "document_name": doc_name,
                    "columns": simplified_columns,
                    "column_meta": column_meta,
                })

                review.processed_documents += 1
                review.results = results
                review.updated_at = utcnow()
                await db.commit()

            # Finalizar
            review.results = results
            review.status = ReviewTableStatus.COMPLETED.value
            review.accuracy_score = (
                round(total_confidence / total_extractions, 3)
                if total_extractions > 0
                else None
            )
            review.updated_at = utcnow()
            await db.commit()
            await db.refresh(review)

            logger.info(
                "Review table concluida: id=%s, docs=%d, accuracy=%.3f",
                review.id,
                review.processed_documents,
                review.accuracy_score or 0,
            )

        except Exception as e:
            logger.error("Erro ao processar review table %s: %s", review_id, e, exc_info=True)
            review.status = ReviewTableStatus.FAILED.value
            review.error_message = str(e)
            review.updated_at = utcnow()
            await db.commit()
            raise

        return review

    # -----------------------------------------------------------------------
    # fill_table — Preenchimento automatico de review table
    # -----------------------------------------------------------------------

    async def fill_table(
        self,
        table_id: str,
        document_ids: Optional[List[str]],
        db: AsyncSession,
    ) -> ReviewTable:
        """Preenche (ou repreenche) a review table com dados extraidos via IA.

        Se document_ids for fornecido, processa apenas esses documentos.
        Caso contrario, processa todos os documentos da review table.

        Diferente de process_review, este metodo pode ser chamado
        incrementalmente para adicionar novos documentos a uma tabela
        ja existente.

        Args:
            table_id: ID da review table.
            document_ids: Lista opcional de IDs de documentos para processar.
            db: Sessao do banco de dados.

        Returns:
            A review table atualizada.
        """
        review = await db.get(ReviewTable, table_id)
        if not review:
            raise ValueError(f"Review table {table_id} nao encontrada")

        template = await db.get(ReviewTableTemplate, review.template_id)
        if not template:
            raise ValueError(f"Template {review.template_id} nao encontrado")

        columns = template.columns or []
        if not columns:
            raise ValueError("Template nao possui colunas definidas")

        # Determinar quais documentos processar
        target_doc_ids = document_ids or review.document_ids
        if not target_doc_ids:
            raise ValueError("Nenhum documento para processar")

        # Se novos documentos foram fornecidos, adicionar a lista de document_ids
        if document_ids:
            existing_ids = set(review.document_ids or [])
            new_ids = [d for d in document_ids if d not in existing_ids]
            if new_ids:
                all_ids = list(review.document_ids or []) + new_ids
                review.document_ids = all_ids
                review.total_documents = len(all_ids)

        # Marcar como processando
        review.status = ReviewTableStatus.PROCESSING.value
        review.error_message = None
        review.updated_at = utcnow()
        await db.commit()

        # Manter resultados existentes para documentos que nao estao sendo reprocessados
        existing_results = {
            r["document_id"]: r
            for r in (review.results or [])
            if r.get("document_id") not in (target_doc_ids or [])
        }

        semaphore = asyncio.Semaphore(MAX_CONCURRENT_EXTRACTIONS)
        new_results: List[Dict[str, Any]] = []
        total_confidence = 0.0
        total_extractions = 0
        processed = 0

        try:
            for doc_id in target_doc_ids:
                doc_result = await db.execute(
                    select(Document).where(Document.id == doc_id)
                )
                doc = doc_result.scalar_one_or_none()

                if not doc:
                    new_results.append({
                        "document_id": doc_id,
                        "document_name": "Documento nao encontrado",
                        "columns": {
                            col["name"]: "Erro: documento nao encontrado"
                            for col in columns
                        },
                    })
                    processed += 1
                    continue

                doc_text = doc.extracted_text or doc.content or ""
                if not doc_text.strip():
                    new_results.append({
                        "document_id": doc_id,
                        "document_name": getattr(doc, "name", doc_id),
                        "columns": {
                            col["name"]: "Erro: sem texto extraido"
                            for col in columns
                        },
                    })
                    processed += 1
                    continue

                doc_text_truncated = doc_text[:MAX_DOC_TEXT_LENGTH]
                doc_name = (
                    getattr(doc, "name", None)
                    or getattr(doc, "title", None)
                    or doc_id
                )

                row_data = await self._extract_row(
                    doc_text=doc_text_truncated,
                    columns=columns,
                    semaphore=semaphore,
                )

                for col_result in row_data.values():
                    if isinstance(col_result, dict) and "confidence" in col_result:
                        total_confidence += col_result["confidence"]
                        total_extractions += 1

                simplified_columns = {}
                column_meta: Dict[str, Any] = {"confidence": {}, "source_excerpt": {}}
                for col_name, col_result in row_data.items():
                    if isinstance(col_result, dict):
                        simplified_columns[col_name] = col_result.get("value", "Nao encontrado")
                        conf = col_result.get("confidence")
                        excerpt = col_result.get("source_excerpt")
                        if conf is not None:
                            try:
                                column_meta["confidence"][col_name] = float(conf)
                            except (TypeError, ValueError):
                                pass
                        if isinstance(excerpt, str) and excerpt.strip():
                            column_meta["source_excerpt"][col_name] = excerpt.strip()
                    else:
                        simplified_columns[col_name] = str(col_result)

                new_results.append({
                    "document_id": doc_id,
                    "document_name": doc_name,
                    "columns": simplified_columns,
                    "column_meta": column_meta,
                })

                processed += 1

                # Atualizar progresso parcial
                all_results = list(existing_results.values()) + new_results
                review.results = all_results
                review.processed_documents = len(all_results)
                review.updated_at = utcnow()
                await db.commit()

            # Finalizar
            final_results = list(existing_results.values()) + new_results
            review.results = final_results
            review.processed_documents = len(final_results)
            review.total_documents = len(review.document_ids or [])
            review.status = ReviewTableStatus.COMPLETED.value
            review.accuracy_score = (
                round(total_confidence / total_extractions, 3)
                if total_extractions > 0
                else None
            )
            review.updated_at = utcnow()
            await db.commit()
            await db.refresh(review)

            logger.info(
                "Fill table concluido: id=%s, novos_docs=%d, total=%d, accuracy=%.3f",
                review.id,
                processed,
                len(final_results),
                review.accuracy_score or 0,
            )

        except Exception as e:
            logger.error("Erro ao preencher review table %s: %s", table_id, e, exc_info=True)
            review.status = ReviewTableStatus.FAILED.value
            review.error_message = str(e)
            review.updated_at = utcnow()
            await db.commit()
            raise

        return review

    # -----------------------------------------------------------------------
    # export_review
    # -----------------------------------------------------------------------

    async def export_review(
        self,
        review_id: str,
        format: str,
        db: AsyncSession,
    ) -> Tuple[bytes, str, str]:
        """Exporta resultados como CSV ou XLSX."""
        review = await db.get(ReviewTable, review_id)
        if not review:
            raise ValueError(f"Review table {review_id} nao encontrada")

        template = await db.get(ReviewTableTemplate, review.template_id)
        if not template:
            raise ValueError(f"Template {review.template_id} nao encontrado")

        columns = template.columns or []
        col_names = [col["name"] for col in columns]

        safe_name = review.name.replace(" ", "_").replace("/", "_")[:50]

        if format == "csv":
            return self._export_csv(review, col_names, safe_name)
        elif format == "xlsx":
            return self._export_xlsx(review, columns, col_names, safe_name)
        else:
            raise ValueError(f"Formato '{format}' nao suportado. Use: csv, xlsx")

    # -----------------------------------------------------------------------
    # Export helpers
    # -----------------------------------------------------------------------

    def _export_csv(
        self,
        review: ReviewTable,
        col_names: List[str],
        safe_name: str,
    ) -> Tuple[bytes, str, str]:
        """Exporta como CSV UTF-8 com BOM."""
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        header = ["Documento"] + col_names
        writer.writerow(header)

        # Rows
        for row in review.results or []:
            values = [row.get("document_name", "")]
            columns_data = row.get("columns", {})
            for col_name in col_names:
                values.append(str(columns_data.get(col_name, "")))
            writer.writerow(values)

        # BOM + content
        content = "\ufeff" + output.getvalue()
        return (
            content.encode("utf-8"),
            f"{safe_name}.csv",
            "text/csv; charset=utf-8",
        )

    def _export_xlsx(
        self,
        review: ReviewTable,
        columns: List[Dict[str, Any]],
        col_names: List[str],
        safe_name: str,
    ) -> Tuple[bytes, str, str]:
        """Exporta como XLSX usando openpyxl com formatacao avancada.

        Inclui:
        - Header estilizado com cores por tipo de coluna
        - Color coding por conteudo da celula
        - Aba de resumo com estatisticas
        - Aba de metadados com definicoes das colunas
        - Freeze panes e auto-width
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValueError(
                "Biblioteca openpyxl nao disponivel. "
                "Instale com: pip install openpyxl"
            )

        wb = Workbook()

        # ----- Aba principal: Dados -----
        ws = wb.active
        ws.title = review.name[:31]  # Excel limita a 31 chars

        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        # Fills por tipo de coluna (header)
        type_header_fills = {
            "risk_rating": PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid"),
            "compliance_check": PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid"),
            "yes_no_classification": PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid"),
            "boolean": PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid"),
        }

        # Fills para color coding de status das celulas
        fill_verified = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_flagged = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        fill_needs_review = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        fill_not_found = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        fill_risk_high = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
        fill_risk_medium = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
        fill_risk_low = PatternFill(start_color="BBF7D0", end_color="BBF7D0", fill_type="solid")
        fill_risk_critical = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

        cell_align = Alignment(vertical="top", wrap_text=True)

        # Montar mapa de tipo por coluna
        col_type_map: Dict[str, str] = {}
        for col_def in columns:
            col_type_map[col_def["name"]] = col_def.get("type", "text")

        # Header
        headers = ["Documento"] + col_names
        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

            # Cor do header baseada no tipo da coluna
            if col_idx > 1:
                col_type = col_type_map.get(header_text, "text")
                cell.fill = type_header_fills.get(col_type, header_fill)
            else:
                cell.fill = header_fill

        # Rows com color coding avancado
        for row_idx, row in enumerate(review.results or [], 2):
            doc_cell = ws.cell(row=row_idx, column=1, value=row.get("document_name", ""))
            doc_cell.font = Font(bold=True, size=10)
            doc_cell.alignment = cell_align
            doc_cell.border = thin_border

            columns_data = row.get("columns", {})
            for col_idx, col_name in enumerate(col_names, 2):
                value = str(columns_data.get(col_name, ""))
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_align
                cell.border = thin_border

                col_type = col_type_map.get(col_name, "text")
                val_lower = value.lower().strip()

                # Color coding por tipo de coluna
                if col_type == "risk_rating":
                    if "critico" in val_lower or "critical" in val_lower:
                        cell.fill = fill_risk_critical
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif "alto" in val_lower or "high" in val_lower:
                        cell.fill = fill_risk_high
                    elif "medio" in val_lower or "medium" in val_lower:
                        cell.fill = fill_risk_medium
                    elif "baixo" in val_lower or "low" in val_lower:
                        cell.fill = fill_risk_low
                    else:
                        cell.fill = fill_not_found
                elif col_type == "compliance_check":
                    if "nao conforme" in val_lower:
                        cell.fill = fill_flagged
                    elif "parcialmente" in val_lower:
                        cell.fill = fill_needs_review
                    elif "conforme" in val_lower:
                        cell.fill = fill_verified
                    else:
                        cell.fill = fill_not_found
                else:
                    # Color coding generico
                    if val_lower in ("nao encontrado", "erro", "erro na extracao", ""):
                        cell.fill = fill_flagged
                    elif val_lower in ("nao", "nao se aplica", "n/a"):
                        cell.fill = fill_needs_review
                    elif val_lower.startswith("erro:"):
                        cell.fill = fill_flagged
                    elif val_lower in ("sim",):
                        cell.fill = fill_verified
                    elif not val_lower:
                        cell.fill = fill_not_found
                    else:
                        cell.fill = fill_verified

        # Auto-width
        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, max(ws.max_row + 1, 2))
            )
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # Freeze panes: fixar header + coluna "Documento"
        ws.freeze_panes = "B2"

        # ----- Aba de Evidencias (trechos/confiança por célula) -----
        evid = wb.create_sheet(title="Evidencias")
        evid_headers = ["Documento", "Coluna", "Trecho", "Confianca"]
        for col_idx, header_text in enumerate(evid_headers, 1):
            cell = evid.cell(row=1, column=col_idx, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        evid_row = 2
        for row in review.results or []:
            doc_name = str(row.get("document_name", "") or "")
            meta = row.get("column_meta") or {}
            excerpts = meta.get("source_excerpt") if isinstance(meta, dict) else {}
            confidences = meta.get("confidence") if isinstance(meta, dict) else {}
            if not isinstance(excerpts, dict):
                excerpts = {}
            if not isinstance(confidences, dict):
                confidences = {}

            for col_name in col_names:
                excerpt = excerpts.get(col_name)
                conf = confidences.get(col_name)
                if not (isinstance(excerpt, str) and excerpt.strip()):
                    continue
                evid.cell(row=evid_row, column=1, value=doc_name).alignment = cell_align
                evid.cell(row=evid_row, column=2, value=col_name).alignment = cell_align
                evid.cell(row=evid_row, column=3, value=str(excerpt).strip()).alignment = cell_align
                try:
                    conf_val = float(conf) if conf is not None else None
                except (TypeError, ValueError):
                    conf_val = None
                if conf_val is not None:
                    evid.cell(row=evid_row, column=4, value=conf_val).alignment = cell_align
                evid_row += 1

        evid.freeze_panes = "A2"
        for col_idx in range(1, len(evid_headers) + 1):
            max_len = max(
                len(str(evid.cell(row=r, column=col_idx).value or ""))
                for r in range(1, max(evid.max_row + 1, 2))
            )
            col_letter = get_column_letter(col_idx)
            evid.column_dimensions[col_letter].width = min(max_len + 4, 80 if col_idx == 3 else 40)

        # ----- Aba de Resumo -----
        ws_summary = wb.create_sheet(title="Resumo")
        summary_header_font = Font(bold=True, size=12, color="1E293B")
        summary_label_font = Font(bold=True, size=10)

        ws_summary.cell(row=1, column=1, value="Resumo da Review Table").font = summary_header_font
        ws_summary.cell(row=3, column=1, value="Nome:").font = summary_label_font
        ws_summary.cell(row=3, column=2, value=review.name)
        ws_summary.cell(row=4, column=1, value="Status:").font = summary_label_font
        ws_summary.cell(row=4, column=2, value=review.status)
        ws_summary.cell(row=5, column=1, value="Documentos processados:").font = summary_label_font
        ws_summary.cell(row=5, column=2, value=f"{review.processed_documents}/{review.total_documents}")
        ws_summary.cell(row=6, column=1, value="Score de confianca:").font = summary_label_font
        ws_summary.cell(
            row=6, column=2,
            value=f"{review.accuracy_score:.1%}" if review.accuracy_score else "N/A",
        )
        ws_summary.cell(row=7, column=1, value="Exportado em:").font = summary_label_font
        ws_summary.cell(row=7, column=2, value=utcnow().strftime("%d/%m/%Y %H:%M UTC"))

        # Estatisticas por coluna
        ws_summary.cell(row=9, column=1, value="Estatisticas por Coluna").font = summary_header_font
        ws_summary.cell(row=10, column=1, value="Coluna").font = summary_label_font
        ws_summary.cell(row=10, column=2, value="Tipo").font = summary_label_font
        ws_summary.cell(row=10, column=3, value="Preenchidos").font = summary_label_font
        ws_summary.cell(row=10, column=4, value="Nao encontrados").font = summary_label_font

        results_data = review.results or []
        for i, col_def in enumerate(columns):
            row_num = 11 + i
            col_n = col_def["name"]
            col_t = col_def.get("type", "text")
            filled = 0
            not_found = 0
            for r in results_data:
                val = r.get("columns", {}).get(col_n, "")
                val_lower = str(val).lower().strip()
                if val_lower in ("nao encontrado", "erro", "", "erro na extracao") or val_lower.startswith("erro:"):
                    not_found += 1
                else:
                    filled += 1

            ws_summary.cell(row=row_num, column=1, value=col_n)
            ws_summary.cell(row=row_num, column=2, value=col_t)
            ws_summary.cell(row=row_num, column=3, value=filled)
            ws_summary.cell(row=row_num, column=4, value=not_found)

        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 25
        ws_summary.column_dimensions["C"].width = 15
        ws_summary.column_dimensions["D"].width = 18

        # ----- Aba de Metadados (definicoes das colunas) -----
        ws_meta = wb.create_sheet(title="Metadados")
        ws_meta.cell(row=1, column=1, value="Definicoes das Colunas").font = summary_header_font
        ws_meta.cell(row=2, column=1, value="Coluna").font = summary_label_font
        ws_meta.cell(row=2, column=2, value="Tipo").font = summary_label_font
        ws_meta.cell(row=2, column=3, value="Prompt de Extracao").font = summary_label_font

        for i, col_def in enumerate(columns):
            row_num = 3 + i
            ws_meta.cell(row=row_num, column=1, value=col_def.get("name", ""))
            ws_meta.cell(row=row_num, column=2, value=col_def.get("type", "text"))
            ws_meta.cell(row=row_num, column=3, value=col_def.get("extraction_prompt", ""))

        ws_meta.column_dimensions["A"].width = 30
        ws_meta.column_dimensions["B"].width = 25
        ws_meta.column_dimensions["C"].width = 80

        # Salvar
        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        buffer.close()

        return (
            content,
            f"{safe_name}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # -----------------------------------------------------------------------
    # Metodos internos
    # -----------------------------------------------------------------------

    async def _extract_row(
        self,
        doc_text: str,
        columns: List[Dict[str, Any]],
        semaphore: asyncio.Semaphore,
    ) -> Dict[str, Any]:
        """Extrai todas as colunas de um documento em paralelo."""

        async def _extract_one(col: Dict[str, Any]) -> Tuple[str, Any]:
            async with semaphore:
                col_name = col["name"]
                col_type = col.get("type", "text")
                extraction_prompt = col.get("extraction_prompt", f"Extraia: {col_name}")

                prompt = EXTRACTION_PROMPT.format(
                    document_text=doc_text,
                    extraction_prompt=extraction_prompt,
                    data_type=COLUMN_TYPE_DESCRIPTIONS.get(col_type, "Texto livre"),
                )

                response = await _call_ai(
                    prompt=prompt,
                    system_instruction=(
                        "Voce e um assistente juridico especializado em extracao "
                        "precisa de dados de documentos. Responda em JSON valido."
                    ),
                    temperature=0.1,
                )

                parsed = _safe_json_parse(response) if response else None

                if parsed and isinstance(parsed, dict):
                    return col_name, {
                        "value": parsed.get("value", "Nao encontrado"),
                        "confidence": min(max(float(parsed.get("confidence", 0.5)), 0.0), 1.0),
                        "source_excerpt": parsed.get("source_excerpt", ""),
                    }

                # Fallback: usar resposta bruta
                return col_name, {
                    "value": response.strip() if response else "Erro na extracao",
                    "confidence": 0.3,
                    "source_excerpt": "",
                }

        tasks = [_extract_one(col) for col in columns]
        results = await asyncio.gather(*tasks, return_exceptions=True)

        row: Dict[str, Any] = {}
        for result in results:
            if isinstance(result, Exception):
                logger.error("Erro na extracao de coluna: %s", result)
                continue
            col_name, col_data = result
            row[col_name] = col_data

        return row

    # -----------------------------------------------------------------------
    # query_review_table — Natural Language Query
    # -----------------------------------------------------------------------

    async def query_review_table(
        self,
        review: ReviewTable,
        question: str,
        db: AsyncSession,
    ) -> Tuple[str, List[Dict[str, Any]]]:
        """Responde perguntas sobre os dados da review table usando LLM."""
        template = await db.get(ReviewTableTemplate, review.template_id)
        col_names = [c["name"] for c in (template.columns if template else [])]

        # Formatar dados da tabela como texto estruturado
        formatted_data = self._format_table_for_query(review.results, col_names)

        # Limitar contexto para evitar exceder janela do modelo
        max_context = 25000
        if len(formatted_data) > max_context:
            formatted_data = formatted_data[:max_context] + "\n\n[... tabela truncada por tamanho]"

        prompt = f"""Voce e um assistente juridico analitico. Analise os dados da review table abaixo e responda a pergunta do usuario.

## DADOS DA REVIEW TABLE: "{review.name}"
### Colunas: {', '.join(col_names)}

{formatted_data}

## PERGUNTA
{question}

## INSTRUCOES
1. Responda de forma clara, objetiva e em portugues brasileiro.
2. Cite os documentos especificos que fundamentam sua resposta.
3. Se a informacao nao estiver disponivel nos dados, diga claramente.
4. Para perguntas quantitativas, faca calculos precisos.
5. Responda em JSON:
```json
{{
  "answer": "<resposta detalhada>",
  "referenced_documents": [
    {{"document_id": "<id>", "document_name": "<nome>", "column_name": "<coluna relevante ou null>"}}
  ]
}}
```

Responda APENAS com o JSON, sem texto adicional."""

        response = await _call_ai(
            prompt=prompt,
            system_instruction=(
                "Voce e um analista juridico especializado em dados tabulares. "
                "Responda com precisao baseando-se exclusivamente nos dados fornecidos."
            ),
            max_tokens=4000,
            temperature=0.2,
        )

        if not response:
            return (
                "Nao foi possivel processar a consulta no momento. Tente novamente.",
                [],
            )

        parsed = _safe_json_parse(response)
        if parsed and isinstance(parsed, dict):
            answer = parsed.get("answer", response)
            refs = parsed.get("referenced_documents", [])
            sources = [
                {
                    "document_id": ref.get("document_id", ""),
                    "document_name": ref.get("document_name", ""),
                    "column_name": ref.get("column_name"),
                }
                for ref in refs
                if isinstance(ref, dict)
            ]
            return answer, sources

        # Fallback: resposta bruta
        return response.strip(), []

    def _format_table_for_query(
        self,
        results: List[Dict[str, Any]],
        col_names: List[str],
    ) -> str:
        """Formata os dados da tabela como texto legivel para o LLM."""
        if not results:
            return "(tabela vazia)"

        lines: List[str] = []
        for i, row in enumerate(results, 1):
            doc_name = row.get("document_name", "Desconhecido")
            doc_id = row.get("document_id", "")
            columns = row.get("columns", {})

            line = f"### Documento {i}: {doc_name} (id: {doc_id})"
            lines.append(line)
            for col in col_names:
                val = columns.get(col, "-")
                lines.append(f"  - {col}: {val}")
            lines.append("")

        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Singleton
# ---------------------------------------------------------------------------

review_table_service = ReviewTableService()
